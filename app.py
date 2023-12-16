import os
from flask import Flask, render_template, request, send_file, jsonify
import pandas as pd
import openpyxl
import tempfile
from flask import Flask, send_from_directory, render_template_string
from datetime import datetime



app = Flask(__name__)

# Initialize a global visit count variable
visit_count = 0

@app.route('/')
def index():
    global visit_count  # Access the global variable
    visit_count += 1  # Increment the visit count
    return render_template('index.html', visit_count=visit_count)


@app.route('/admin')
def admin():
    # Add authentication here to protect the admin page
    # For simplicity, you can just check a password (not recommended for production)
    password = request.args.get('password')
    if password == 'lyinginwait':
        return render_template('admin.html')
    else:
        return 'Unauthorized'

# Create a route to get the visit count as JSON
@app.route('/get_visit_count')
def get_visit_count():
    return jsonify(visit_count=visit_count)




@app.route('/ELEC.png')
def favicon():
    return send_from_directory('static', 'ELEC.png', mimetype='image/x-icon')


@app.route('/download_template')
def download_template():
    template_path = 'static/GST_PORTAL_TEMPLATE.xlsx'  # Adjust the path as needed
    return send_from_directory('', template_path, as_attachment=True)




@app.route('/compare', methods=['POST'])
def compare():
    gst_file = request.files['gst_file']
    erp_file = request.files['erp_file']

    if not gst_file or not erp_file:
        return "Please upload both GST and ERP files."

    # Load GST file into a Pandas DataFrame
    gst_df = pd.read_excel(gst_file,  header=5)
    gst_df['Invoice number'] = gst_df['Invoice number'].str.replace("'", '', regex=False)

    grouped_gst_df = gst_df.groupby(['Invoice number', 'Supplier GSTIN'], as_index=False).agg({
        'Supplier ID': 'first',
        'IGST': 'sum',
        'CGST': 'sum',
        'SGST': 'sum',
        'Taxable Amount':'sum'


    })



    gst_df = grouped_gst_df


    gst_df = gst_df[['Supplier ID', 'Invoice number', 'IGST', 'SGST', 'CGST', 'Taxable Amount', 'Supplier GSTIN']]

    # Load ERP file into another Pandas DataFrame
    erp_df = pd.read_excel(erp_file,  header=0)  # Adjust the header as needed

    # Process ERP data as per your code
    erp_df['IGST_ERP'] = erp_df['Input IGST 0% - EIPL'] + erp_df['Input IGST 12% - EIPL'] + erp_df[
        'Input IGST 18% - EIPL'] + erp_df['Input IGST 28% - EIPL'] + erp_df['Input IGST 5% - EIPL']

    erp_df['SGST_ERP'] = erp_df['Input SGST 0% - EIPL'] + erp_df['Input SGST 14% - EIPL'] + erp_df[
        'Input SGST 2.5% - EIPL'] + erp_df['Input SGST 6% - EIPL'] + erp_df['Input SGST 9% - EIPL']

    erp_df['CGST_ERP'] = erp_df['Input CGST 0% - EIPL'] + erp_df['Input CGST 14% - EIPL'] + erp_df[
        'Input CGST 2.5% - EIPL'] + erp_df['Input CGST 6% - EIPL'] + erp_df['Input CGST 9% - EIPL']

    erp_df = erp_df[
        [ 'Bill Date','Supplier Id', 'Bill No','Supplier GSTIN', 'IGST_ERP',
         'SGST_ERP', 'CGST_ERP', 'Net Total']]

    # Assuming you have NaN values for empty 'Supplier Id'
    erp_df.dropna(subset=['Supplier Id'], inplace=True)

    # If you want to reset the index after dropping rows
    erp_df.reset_index(drop=True, inplace=True)


    # Assuming 'erp_df' is your DataFrame
    # erp_df['Taxable'] = erp_df['Net Total'] + erp_df['Freight & Forwarding - Purchase - EIPL']
    erp_df.rename(columns={'Net Total': 'Taxable'}, inplace=True)  # Use inplace=True to modify the DataFrame in-place



    # Merge the two DataFrames based on the GST number
    #filtered_df = pd.merge(gst_df, erp_df, on='Supplier GSTIN')

    # Merge the two DataFrames based on the GST number
    merged_df = pd.merge(gst_df, erp_df, on='Supplier GSTIN')

    # Filter rows as per your code
    filtered_df = merged_df[merged_df.apply(
        lambda x: ' '.join(str(x['Invoice number']).split()[-3:]) == ' '.join(str(x['Bill No']).split()[-3:]), axis=1)]

    # Create a copy of the original DataFrame
    filtered_df_gt_10 = filtered_df.copy()
    filtered_df_lt_10 = filtered_df.copy()

    # Filter rows based on the condition
    filtered_df_gt_10 = filtered_df_gt_10[
        (abs(filtered_df_gt_10['CGST'] - filtered_df_gt_10['CGST_ERP']) > 100) |
        (abs(filtered_df_gt_10['IGST'] - filtered_df_gt_10['IGST_ERP']) > 100) |
        (abs(filtered_df_gt_10['SGST'] - filtered_df_gt_10['SGST_ERP']) > 100) |
        (abs(filtered_df_gt_10['Taxable Amount'] - filtered_df_gt_10['Taxable']) > 100)
        ]

    filtered_df_lt_10 = filtered_df_lt_10[
        (abs(filtered_df_lt_10['CGST'] - filtered_df_lt_10['CGST_ERP']) <= 100) &
        (abs(filtered_df_lt_10['IGST'] - filtered_df_lt_10['IGST_ERP']) <= 100) &
        (abs(filtered_df_lt_10['SGST'] - filtered_df_lt_10['SGST_ERP']) <= 100) &
        (abs(filtered_df_lt_10['Taxable Amount'] - filtered_df_lt_10['Taxable']) <= 100)
        ]
    # Assuming 'filtered_df_gt_10' is your DataFrame
    filtered_df_gt_10['CGST Difference'] = abs(filtered_df_gt_10['CGST'] - filtered_df_gt_10['CGST_ERP'])
    filtered_df_gt_10['IGST Difference'] = abs(filtered_df_gt_10['IGST'] - filtered_df_gt_10['IGST_ERP'])
    filtered_df_gt_10['SGST Difference'] = abs(filtered_df_gt_10['SGST'] - filtered_df_gt_10['SGST_ERP'])
    filtered_df_gt_10['Taxable Amount Difference'] = abs(
        filtered_df_gt_10['Taxable Amount'] - filtered_df_gt_10['Taxable'])
    # Extract the 'Supplier GSTIN' values from the DataFrames
    filtered_df_gt_10_gstin = filtered_df_gt_10['Supplier GSTIN'].tolist()
    filtered_df_lt_10_gstin = filtered_df_lt_10['Supplier GSTIN'].tolist()

    # Create a copy of the original DataFrames
    df_copy = gst_df.copy()

    erp_df_copy = erp_df.copy()

    # Merge the two DataFrames based on the 'Supplier GSTIN' column
    merged_data = pd.merge(df_copy, erp_df_copy, on='Supplier GSTIN', how='outer', indicator=True)

    # Rows that are in 'df' but not in 'erp_df'
    left_only = merged_data[merged_data['_merge'] == 'left_only'].copy()
    left_only.drop(columns=['_merge'], inplace=True)

    # Rows that are in 'erp_df' but not in 'df'
    right_only = merged_data[merged_data['_merge'] == 'right_only'].copy()
    right_only.drop(columns=['_merge'], inplace=True)

    # Create the comparison Excel file
    comparison_filename = os.path.join(tempfile.gettempdir(), 'comparison.xlsx')
    with pd.ExcelWriter(comparison_filename, engine='openpyxl') as writer:
        # Write filtered_df_gt_10 to the first sheet
        filtered_df_gt_10.to_excel(writer, sheet_name='Unmatched', index=False)

        # Write filtered_df_lt_10 to the second sheet
        filtered_df_lt_10.to_excel(writer, sheet_name='Matched', index=False)

        # Write left_only to the third sheet
        left_only.to_excel(writer, sheet_name='Non-existence ERP Portal', index=False)

        # Write right_only to the fourth sheet
        right_only.to_excel(writer, sheet_name='Non-Existence GST Portal', index=False)

        # Access the Excel writer workbook and worksheets
        workbook = writer.book
        worksheet_gt_10 = writer.sheets['Unmatched']
        worksheet_lt_10 = writer.sheets['Matched']
        worksheet_unmatched_left = writer.sheets['Non-existence ERP Portal']
        worksheet_unmatched_right = writer.sheets['Non-Existence GST Portal']

        # Define a style for columns I to O (light blue) for both sheets
        light_blue_fill = openpyxl.styles.PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')



        # Apply light blue fill to columns I to O for both sheets
        for worksheet in [worksheet_gt_10, worksheet_lt_10, worksheet_unmatched_left, worksheet_unmatched_right]:
            for col in worksheet.iter_cols(min_col=9, max_col=13):
                for cell in col:
                    cell.fill = light_blue_fill

        light_red_fill = openpyxl.styles.PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')

        for worksheet in [worksheet_gt_10, worksheet_lt_10, worksheet_unmatched_left, worksheet_unmatched_right]:
            for col in worksheet.iter_cols(min_col=14, max_col=18):
                for cell in col:
                    cell.fill = light_red_fill

        # Define a style for cells with differences greater than 10 (yellow) in the "Filtered Data > 10" sheet
        teel_fill = openpyxl.styles.PatternFill(start_color='008080', end_color='008080', fill_type='solid')

        # Apply yellow fill to cells with differences greater than 10 in the "Filtered Data > 10" sheet
        for worksheet in [worksheet_gt_10, worksheet_lt_10, worksheet_unmatched_left, worksheet_unmatched_right]:
            for col in worksheet.iter_cols(min_col=1, max_col=6):
                for cell in col:
                    cell.fill = teel_fill

        white_fill = openpyxl.styles.PatternFill(start_color='FAF9F6', end_color='FAF9F6', fill_type='solid')
        for worksheet in [worksheet_gt_10, worksheet_lt_10, worksheet_unmatched_left, worksheet_unmatched_right]:
            for col in worksheet.iter_cols(min_col=7, max_col=8):
                for cell in col:
                    cell.fill = white_fill




        # Auto-adjust column widths for both sheets
        for worksheet in [worksheet_gt_10, worksheet_lt_10, worksheet_unmatched_left, worksheet_unmatched_right]:
            for column_cells in worksheet.columns:
                max_length = 0
                column = column_cells[0].column_letter  # Get the column name
                for cell in column_cells:
                    try:  # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column].width = adjusted_width

        # Get the current date in a specific format (e.g., YYYY-MM-DD)
        current_date = datetime.now().strftime('%Y-%m-%d')

        # Define the desired file name
        download_name = f'{current_date}_GST_comparison.xlsx'


    return send_file(comparison_filename, as_attachment=True, download_name= download_name)

if __name__ == '__main__':
    app.run(host="0.0.0.0", port=8000)
